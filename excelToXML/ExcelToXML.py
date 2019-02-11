from openpyxl import load_workbook
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from xml.dom.minidom import parse, parseString

wb = load_workbook(filename = "ExcelBook.xlsx")
ws = wb['TestSheet']

greet = []

class Greetings:
    def __init__ (self, text, number, desc, index):
        self.text = text
        self.number = number
        self.desc = desc
        self.index = index

def setGreetings():
    for row in ws.iter_rows(min_row = 3, max_col = 4, max_row = 10):
        temp = Greetings(row[0].value, row[1].value, row[2].value, row[3].value)
        greet.append(temp)

def main():
    setGreetings()
    root = ET.Element('Greetings')
    for greets in greet:
        base = ET.SubElement(root, 'Greeting')
        baseText = ET.SubElement(base, 'Text')
        baseText.text = str(greets.text)
        baseDesc = ET.SubElement(base, 'Description')
        baseDesc.text = str(greets.desc)
        baseIndex = ET.SubElement(base, "Index")
        baseIndex.text = str(greets.index)
        baseNumber = ET.SubElement(base, "Number")
        baseNumber.text = str(greets.number)
    myData = ET.ElementTree(root)
    myData.write("text.html")

main()